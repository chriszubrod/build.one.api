# Python Standard Library Imports
import logging
from collections import defaultdict
from copy import copy
from io import BytesIO
from typing import Any, Dict, List, Optional

# Third-party Imports
# openpyxl is imported lazily inside apply_rows_to_details so this module
# imports (and py_compiles) even in environments where openpyxl is not yet
# installed (it is in requirements.txt at ~3.1.2, present in prod/CI).

logger = logging.getLogger(__name__)


# DETAILS row shape (26 cols, A..Z). Ported verbatim from the MS Graph Excel
# sync (entities/bill/business/service.py::sync_to_excel_workbook). Indices:
#   0  A empty
#   1  B CostCode number
#   2  C SubCostCode number   <- grouping/insertion key
#   3..7 D-H empty
#   8  I date
#   9  J vendor name
#   10 K number (bill/expense/credit number)
#   11 L description
#   12 M type label ("Bill" | "Expense" | "Credit")
#   13 N price/amount (Decimal-safe)
#   14..24 O-Y empty
#   25 Z line-item public_id (idempotency / reconciliation key)
DETAILS_ROW_WIDTH = 26
SUBCOSTCODE_COL_INDEX = 2  # column C (0-based)
DATE_COL_INDEX = 8  # column I (0-based) — the date
DEFAULT_KEY_COL_INDEX = 25  # column Z (0-based)

# The date column is written as a real date VALUE (not a text string) so the
# number format actually renders it, with an explicit mm/dd/yyyy display.
DATE_NUMBER_FORMAT = "mm/dd/yyyy"


def _normalize_subcostcode(value: Any) -> str:
    """
    Normalize a SubCostCode cell value for matching: exact string compare with
    a numeric-equivalence fallback (mirrors find_insertion_row_for_subcostcode
    so '65.03' / 65.03 / '65.030' compare consistently).
    """
    if value is None:
        return ""
    return str(value).strip()


def _subcostcode_matches(cell_value: Any, target: str) -> bool:
    """Exact-string-or-numeric-equivalence match, ported from the MS path."""
    cell_str = _normalize_subcostcode(cell_value)
    target_str = _normalize_subcostcode(target)
    if not cell_str or not target_str:
        return False
    if cell_str == target_str:
        return True
    try:
        return float(cell_str) == float(target_str)
    except (ValueError, TypeError):
        return False


def _find_insertion_row(
    ws,
    target_subcostcode: str,
    *,
    max_existing_row: int,
) -> Optional[int]:
    """
    openpyxl port of find_insertion_row_for_subcostcode.

    Operates on the worksheet's pre-existing rows (1..max_existing_row), so the
    insertion points for every group are computed against the ORIGINAL sheet —
    not against rows we just inserted this pass. Returns a 1-based row number to
    insert BEFORE (openpyxl ws.insert_rows(idx) shifts idx downward), or None to
    append at the end.

    Logic (verbatim behavior):
      1. Collect the matching rows (col C == target) and whether each has BOTH
         Date (col I) AND Vendor (col J).
      2. If a matching row has data, insert AFTER the last such row.
      3. Else (matches but no data rows) insert two rows after the first match.
      4. No match at all → None (append at end).

    Row 1 is treated as the header (skipped), mirroring the MS path which
    `continue`s on row_index == 0.
    """
    matching_rows = []  # list of (excel_row, has_date_and_vendor)

    for excel_row in range(2, max_existing_row + 1):
        col_c = ws.cell(row=excel_row, column=SUBCOSTCODE_COL_INDEX + 1).value
        if not _subcostcode_matches(col_c, target_subcostcode):
            continue
        col_i = ws.cell(row=excel_row, column=9).value   # I
        col_j = ws.cell(row=excel_row, column=10).value  # J
        has_data = (
            col_i is not None and str(col_i).strip() != ""
            and col_j is not None and str(col_j).strip() != ""
        )
        matching_rows.append((excel_row, has_data))

    if not matching_rows:
        return None

    last_data_row = None
    for excel_row, has_data in matching_rows:
        if has_data:
            last_data_row = excel_row

    if last_data_row is not None:
        return last_data_row + 1

    return matching_rows[0][0] + 2


# Excel error literals. A `<definedName>` whose VALUE is one of these (e.g. a
# broken `_xlnm.Print_Titles` = `#N/A`) is already non-functional in Excel, but
# openpyxl's strict parser REJECTS THE WHOLE WORKBOOK over it
# (PrintTitles.from_string("#N/A") -> ValueError). Real workbooks accumulate
# these (the OHR2 Budget Tracker had 35). We strip only error-valued defined
# names before load — valid named ranges are untouched, so no formula that
# references a real name can break (a formula pointing at a #REF! name is
# already broken anyway).
_EXCEL_ERROR_LITERALS = ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def _sanitize_workbook_bytes(file_bytes: bytes) -> bytes:
    """
    Return `file_bytes` with error-valued `<definedName>` elements stripped from
    xl/workbook.xml, so openpyxl can load workbooks that Excel tolerates but
    openpyxl rejects. Best-effort: any failure (not a zip / no workbook.xml /
    no defined names) returns the original bytes unchanged.
    """
    import re
    import zipfile

    try:
        src = BytesIO(file_bytes)
        with zipfile.ZipFile(src) as zin:
            if "xl/workbook.xml" not in zin.namelist():
                return file_bytes
            wbxml = zin.read("xl/workbook.xml").decode("utf-8")
            if "<definedName" not in wbxml:
                return file_bytes

            def _drop_if_error(match: "re.Match") -> str:
                inner = match.group(2)
                return "" if any(e in inner for e in _EXCEL_ERROR_LITERALS) else match.group(0)

            cleaned = re.sub(
                r"<definedName\b([^>]*)>(.*?)</definedName>",
                _drop_if_error,
                wbxml,
                flags=re.DOTALL,
            )
            if cleaned == wbxml:
                return file_bytes  # nothing error-valued — leave the file byte-identical

            out = BytesIO()
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    payload = cleaned.encode("utf-8") if item.filename == "xl/workbook.xml" else zin.read(item.filename)
                    zout.writestr(item, payload)
            logger.info(
                "box.excel.workbook_sanitized",
                extra={"event_name": "box.excel.workbook_sanitized"},
            )
            return out.getvalue()
    except Exception as error:
        logger.warning(
            "box.excel.sanitize_skipped",
            extra={"event_name": "box.excel.sanitize_skipped", "error_class": type(error).__name__},
        )
        return file_bytes


def _cell_value_for_write(value: Any) -> Any:
    """
    Coerce a row element into something openpyxl can write without precision
    loss. None → empty cell. Decimal → float (openpyxl serializes Decimal as a
    string otherwise, which breaks the formula summary tabs that SUM column N);
    str(Decimal) preserves the exact digits before the float conversion, so we
    never round-trip through a lossy intermediate.

    int/float/str pass through unchanged.
    """
    if value is None:
        return None
    # Decimal: import locally to avoid a hard module dependency at import time.
    from decimal import Decimal

    if isinstance(value, Decimal):
        # str() preserves the exact decimal digits; float() is what Excel/
        # openpyxl stores numerically so the summary SUM() formulas see a
        # number, not text.
        return float(str(value))
    return value


def _coerce_date(value: Any) -> Any:
    """
    Coerce a date-like value to a real datetime so a date number format renders
    it (a text string like "2026-06-15" would display literally — the format
    only applies to actual date values). date/datetime pass through; common
    string forms are parsed; anything unparseable is returned unchanged so we
    never crash or corrupt a cell.
    """
    from datetime import date, datetime

    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return value
    return value


def _write_row_values(ws, target_row: int, row: List[Any]) -> None:
    """
    Write a 26-col row's values into `target_row`. The date column (I) is
    written as a real date value with the mm/dd/yyyy display format (overriding
    whatever the styling template carried for that column); all other columns
    pass through `_cell_value_for_write`.
    """
    from datetime import datetime

    for col_index, value in enumerate(row):
        cell = ws.cell(row=target_row, column=col_index + 1)
        if col_index == DATE_COL_INDEX:
            coerced = _coerce_date(value)
            cell.value = coerced
            if isinstance(coerced, datetime):
                cell.number_format = DATE_NUMBER_FORMAT
        else:
            cell.value = _cell_value_for_write(value)


# Column probed to recognize a "real, formatted data row" when choosing a
# styling template — N (amount, 1-based 14) reliably carries a currency format
# on populated cost lines; a blank/separator row shows 'General'.
_STYLE_PROBE_COL = 14


def _find_template_row(ws, anchor_row: int, *, limit: int = 200) -> Optional[int]:
    """
    Nearest row at/above `anchor_row` that looks like a real formatted data row
    (its amount cell carries a non-'General' number format), to copy styling
    from. For an insert, `anchor_row` is the matched group's last row (already
    formatted, returned immediately). For an append, we may walk up past blank
    trailing rows. Returns `anchor_row` as a fallback, or None if out of range.
    """
    floor = max(1, anchor_row - limit)
    r = anchor_row
    while r >= floor:
        cell = ws.cell(row=r, column=_STYLE_PROBE_COL)
        if cell.has_style and cell.number_format not in (None, "General"):
            return r
        r -= 1
    return anchor_row if anchor_row >= 1 else None


def _copy_row_style(ws, src_row: Optional[int], dst_row: int, max_col: int) -> None:
    """
    Copy cell formatting (number format, font, border, fill, alignment,
    protection) + row height from `src_row` to `dst_row`. openpyxl inserts BARE
    rows (default Calibri / 'General'), unlike Excel which inherits the row
    above — so without this, every inserted cost line loses its currency / date
    formatting and font. Style objects are copy()'d (openpyxl forbids sharing).
    """
    if not src_row or src_row < 1 or src_row == dst_row:
        return
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        if not src.has_style:
            continue
        dst = ws.cell(row=dst_row, column=col)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format  # str — no copy needed
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    if src_row in ws.row_dimensions and ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def apply_rows_to_details(
    file_bytes: bytes,
    worksheet_name: str,
    rows: List[List[Any]],
    key_col_index: int = DEFAULT_KEY_COL_INDEX,
) -> Dict[str, Any]:
    """
    Apply new DETAILS rows to a Box-hosted .xlsx, preserving formula fidelity.

    Pure / deterministic — no I/O beyond the in-memory BytesIO round trip. This
    is the openpyxl analogue of the MS Graph Excel sync, but Box has no
    cell-level API, so the whole workbook is edited here and the caller uploads
    a new version.

    Contract:
      - `rows` is a list of 26-element lists (Z = line-item public_id at
        `key_col_index`, default 25).
      - Load WITHOUT data_only (data_only=False) so the formula-driven summary
        tabs keep their formulas (NEVER data_only=True — that would replace
        them with stale cached values).
      - Set fullCalcOnLoad=True so Excel / Excel-for-web recalculates the
        summary tabs when next opened.
      - Only ever mutate the DETAILS worksheet (`worksheet_name`).
      - Idempotency: read existing col-Z keys; skip any row whose key is already
        present. If ALL rows are already present → no-op: return bytes=None so
        the caller skips the upload entirely (this also closes the
        crash-after-upload-before-mark_done window).
      - Insertion: group new rows by SubCostCode (col C); for each group find
        the insertion point via the ported find_insertion logic; ws.insert_rows
        + write cells.

    Returns {"bytes": <bytes>|None, "applied": int, "skipped": int}.
    bytes is None iff applied == 0 (all keys already present).
    """
    from openpyxl import load_workbook

    # Real workbooks carry openpyxl-hostile cruft (error-valued print-title /
    # defined names) that Excel tolerates but openpyxl refuses to load. Strip it
    # first; a clean workbook is returned byte-identical (no-op).
    file_bytes = _sanitize_workbook_bytes(file_bytes)
    wb = load_workbook(BytesIO(file_bytes), data_only=False)

    # fullCalcOnLoad: the builder must use whatever attribute the installed
    # openpyxl exposes. 3.1.x exposes wb.calculation.fullCalcOnLoad
    # (CalcProperties). Set defensively so a future openpyxl rename doesn't
    # silently drop the flag without surfacing.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        logger.warning(
            "box.excel.editor.fullcalc_unavailable",
            extra={"event_name": "box.excel.editor.fullcalc_unavailable"},
        )

    if worksheet_name not in wb.sheetnames:
        raise ValueError(
            f"worksheet {worksheet_name!r} not found in workbook "
            f"(tabs: {wb.sheetnames})"
        )
    ws = wb[worksheet_name]

    # 1-based column for the key (Z = 26 when key_col_index = 25).
    key_col = key_col_index + 1

    # The number of ORIGINAL data rows — insertion points are computed against
    # this fixed extent so groups don't see each other's freshly-inserted rows.
    max_existing_row = ws.max_row

    # Idempotency: collect existing col-Z keys.
    existing_keys = set()
    for excel_row in range(1, max_existing_row + 1):
        val = ws.cell(row=excel_row, column=key_col).value
        if val is not None and str(val).strip():
            existing_keys.add(str(val).strip())

    # Filter out rows whose key is already present; preserve input order.
    new_rows = []
    skipped = 0
    for row in rows:
        key = ""
        if len(row) > key_col_index and row[key_col_index] is not None:
            key = str(row[key_col_index]).strip()
        if key and key in existing_keys:
            skipped += 1
            continue
        # Guard against a duplicate key within the same batch.
        if key:
            existing_keys.add(key)
        new_rows.append(row)

    if not new_rows:
        # No-op: all keys already present. Returning bytes=None tells the caller
        # to skip the upload (and makes a retry after a crash a clean no-op).
        return {"bytes": None, "applied": 0, "skipped": skipped}

    # Group new rows by SubCostCode (col C). Use a stable dict so dispatch order
    # is deterministic.
    by_subcostcode: "defaultdict[str, list]" = defaultdict(list)
    group_order: List[str] = []
    for row in new_rows:
        scc = ""
        if len(row) > SUBCOSTCODE_COL_INDEX and row[SUBCOSTCODE_COL_INDEX] is not None:
            scc = str(row[SUBCOSTCODE_COL_INDEX]).strip()
        if scc not in by_subcostcode:
            group_order.append(scc)
        by_subcostcode[scc].append(row)

    # Compute every group's insertion row against the ORIGINAL sheet first.
    # Groups that match get an insertion index; groups with no match (or no
    # SubCostCode) append at the end. Computing all indices up front against the
    # fixed `max_existing_row` keeps them stable; we then apply insertions in
    # descending row order so an earlier insert can't shift a later index.
    planned_inserts = []  # list of (insertion_row, group_rows)
    append_rows: List[List[Any]] = []
    for scc in group_order:
        group_rows = by_subcostcode[scc]
        insertion_row = None
        if scc:
            insertion_row = _find_insertion_row(
                ws, scc, max_existing_row=max_existing_row
            )
        if insertion_row is not None:
            planned_inserts.append((insertion_row, group_rows))
        else:
            append_rows.extend(group_rows)

    applied = 0
    # Columns to carry styling across (insert_rows doesn't change column count).
    style_max_col = ws.max_column

    # Apply inserts top-of-sheet-last: sort by insertion_row DESCENDING so each
    # ws.insert_rows shifts only rows below it — never invalidating a
    # not-yet-applied (smaller) insertion index.
    planned_inserts.sort(key=lambda g: g[0], reverse=True)
    for insertion_row, group_rows in planned_inserts:
        ws.insert_rows(insertion_row, amount=len(group_rows))
        # The row directly above the insertion is the matched group's last
        # (formatted) row — our styling template so the new cost lines inherit
        # the currency/date number formats, font, borders + fill.
        template_row = _find_template_row(ws, insertion_row - 1)
        for offset, row in enumerate(group_rows):
            target_row = insertion_row + offset
            _copy_row_style(ws, template_row, target_row, style_max_col)
            _write_row_values(ws, target_row, row)
            applied += 1

    # Append any groups that found no SubCostCode match at the end of the sheet.
    if append_rows:
        # Recompute the end after the inserts above (insert_rows updated max_row).
        append_start = ws.max_row + 1
        template_row = _find_template_row(ws, append_start - 1)
        for offset, row in enumerate(append_rows):
            target_row = append_start + offset
            _copy_row_style(ws, template_row, target_row, style_max_col)
            _write_row_values(ws, target_row, row)
            applied += 1

    out = BytesIO()
    wb.save(out)
    return {"bytes": out.getvalue(), "applied": applied, "skipped": skipped}


# ---------------------------------------------------------------------------
# Self-test. Run `python integrations/box/excel/business/workbook_editor.py`.
# Builds a tiny in-memory .xlsx with a DETAILS data sheet + a formula summary
# sheet, applies rows, reloads, and asserts:
#   - data rows landed,
#   - the summary formula STRING survived (not converted to a value),
#   - col-Z keys are present,
#   - re-apply is a no-op (bytes is None).
# ---------------------------------------------------------------------------
def _self_test() -> None:  # pragma: no cover - manual harness
    from decimal import Decimal
    from io import BytesIO as _BytesIO

    from openpyxl import Workbook, load_workbook

    # Build a workbook: DETAILS data sheet + Summary formula sheet.
    wb = Workbook()
    details = wb.active
    details.title = "DETAILS"
    # Header row.
    details.append(["", "Cost", "SubCost"] + [""] * 5 + ["Date", "Vendor", "Num", "Desc", "Type", "Price"] + [""] * 11 + ["Key"])
    # One existing data row for SubCostCode 65.03 with date + vendor.
    existing = [""] * 26
    existing[1] = "65"
    existing[2] = "65.03"
    existing[8] = "2026-06-01"
    existing[9] = "Acme"
    existing[10] = "INV-1"
    existing[11] = "old line"
    existing[12] = "Bill"
    existing[13] = 100.0
    existing[25] = "existing-key-aaaa"
    details.append(existing)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total"
    # A SUM over the DETAILS price column — must survive as a formula string.
    summary["B1"] = "=SUM(DETAILS!N:N)"

    buf = _BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    def _make_row(scc, num, desc, price, key):
        r = [""] * 26
        r[1] = scc.split(".")[0]
        r[2] = scc
        r[8] = "2026-06-10"
        r[9] = "Acme"
        r[10] = num
        r[11] = desc
        r[12] = "Bill"
        r[13] = Decimal(price)
        r[25] = key
        return r

    rows = [
        _make_row("65.03", "INV-2", "new matching line", "250.55", "new-key-1111"),
        _make_row("99.99", "INV-3", "no-match append line", "33.33", "new-key-2222"),
        # Duplicate of the already-present key — must be skipped.
        [*( [""] * 25 ), "existing-key-aaaa"],
    ]

    result = apply_rows_to_details(file_bytes, "DETAILS", rows)
    assert result["bytes"] is not None, "expected bytes on a non-empty apply"
    assert result["applied"] == 2, f"expected 2 applied, got {result['applied']}"
    assert result["skipped"] == 1, f"expected 1 skipped, got {result['skipped']}"

    # Reload and verify.
    wb2 = load_workbook(_BytesIO(result["bytes"]), data_only=False)
    d2 = wb2["DETAILS"]
    keys = set()
    prices_seen = {}
    for r in range(1, d2.max_row + 1):
        k = d2.cell(row=r, column=26).value
        if k:
            keys.add(str(k).strip())
        n = d2.cell(row=r, column=14).value
        desc = d2.cell(row=r, column=12).value
        if desc:
            prices_seen[str(desc)] = n
    assert "new-key-1111" in keys, "matching-line key missing"
    assert "new-key-2222" in keys, "append-line key missing"
    assert "existing-key-aaaa" in keys, "existing key disappeared"

    # Decimal price landed as a number, not a string, with no precision loss.
    assert prices_seen.get("new matching line") == 250.55, (
        f"price precision lost: {prices_seen.get('new matching line')!r}"
    )

    # Formula string survived (NOT converted to a cached value).
    s2 = wb2["Summary"]
    assert s2["B1"].value == "=SUM(DETAILS!N:N)", (
        f"summary formula did not survive: {s2['B1'].value!r}"
    )

    # fullCalcOnLoad serialized into calcPr.
    assert wb2.calculation.fullCalcOnLoad is True, "fullCalcOnLoad did not serialize"

    # Re-apply the SAME rows → no-op (all keys now present).
    reapply = apply_rows_to_details(result["bytes"], "DETAILS", rows)
    assert reapply["bytes"] is None, "re-apply should be a no-op (bytes=None)"
    assert reapply["applied"] == 0, "re-apply should apply nothing"

    print("workbook_editor self-test PASSED")


if __name__ == "__main__":  # pragma: no cover
    _self_test()
