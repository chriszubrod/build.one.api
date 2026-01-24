# Python Standard Library Imports
import io
import logging
import re
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional, Dict, List, Tuple, Any

# Third-party Imports
from openpyxl import load_workbook
import xlrd  # For .xls files

# Local Imports
from modules.contract_labor.business.model import ContractLabor
from modules.contract_labor.persistence.repo import ContractLaborRepository
from modules.vendor.business.service import VendorService
from modules.project.business.service import ProjectService

logger = logging.getLogger(__name__)


class ContractLaborImportService:
    """
    Service for importing contract labor time logs from Excel files.
    
    Excel Format Expected:
    - Column A: Date (e.g., "Tuesday, January 20, 2026")
    - Column B: Job (Project name/abbreviation like "HP - 6135 Hillsboro Pike")
    - Column C: Name (Vendor/contractor name)
    - Column D: Time In
    - Column E: Time Out
    - Column F: Break Time
    - Column G: Regular Time (e.g., "02:45" or "08:30")
    - Column H: OT (Overtime)
    - Column I: Total Work Time
    - Column J: Notes
    """

    # Excel column indices (0-based)
    COL_DATE = 0          # A
    COL_JOB = 1           # B
    COL_NAME = 2          # C
    COL_TIME_IN = 3       # D
    COL_TIME_OUT = 4      # E
    COL_BREAK_TIME = 5    # F
    COL_REGULAR_TIME = 6  # G
    COL_OT = 7            # H
    COL_TOTAL_TIME = 8    # I
    COL_NOTES = 9         # J

    def __init__(self):
        """Initialize the import service."""
        self.repo = ContractLaborRepository()
        self.vendor_service = VendorService()
        self.project_service = ProjectService()
        
        # Cache for vendor and project lookups
        self._vendor_cache: Dict[str, Any] = {}
        self._project_cache: Dict[str, Any] = {}
        self._rate_cache: Dict[int, Tuple[Optional[Decimal], Optional[Decimal]]] = {}

    def _load_excel_rows(self, file_content: bytes, filename: str) -> Optional[List[Tuple]]:
        """
        Load Excel file and return rows as list of tuples.
        Supports both .xlsx (openpyxl) and .xls (xlrd) formats.
        
        Args:
            file_content: Excel file content as bytes
            filename: Original filename to detect format
            
        Returns:
            List of row tuples (skipping header), or None on error
        """
        filename_lower = filename.lower()
        
        try:
            if filename_lower.endswith('.xlsx'):
                # Use openpyxl for .xlsx files
                workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
                sheet = workbook.active
                if not sheet:
                    logger.error("No active sheet found in .xlsx file")
                    return None
                # Return rows starting from row 2 (skip header)
                return [tuple(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
            
            elif filename_lower.endswith('.xls'):
                # Use xlrd for .xls files
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                if sheet.nrows < 2:
                    logger.warning("Excel file has no data rows")
                    return []
                # Return rows starting from row 1 (skip header at row 0)
                rows = []
                for row_idx in range(1, sheet.nrows):
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        # Handle xlrd date cells
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                value = xlrd.xldate_as_datetime(value, workbook.datemode)
                            except Exception:
                                pass
                        row_values.append(value)
                    rows.append(tuple(row_values))
                return rows
            
            else:
                logger.error(f"Unsupported file format: {filename}")
                return None
                
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            # Try the other format as fallback
            try:
                if filename_lower.endswith('.xlsx'):
                    # .xlsx failed, try as .xls
                    workbook = xlrd.open_workbook(file_contents=file_content)
                    sheet = workbook.sheet_by_index(0)
                    rows = []
                    for row_idx in range(1, sheet.nrows):
                        row_values = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            value = cell.value
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    value = xlrd.xldate_as_datetime(value, workbook.datemode)
                                except Exception:
                                    pass
                            row_values.append(value)
                        rows.append(tuple(row_values))
                    logger.info(f"Successfully loaded {filename} as .xls format")
                    return rows
                else:
                    # .xls failed, try as .xlsx
                    workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
                    sheet = workbook.active
                    if sheet:
                        logger.info(f"Successfully loaded {filename} as .xlsx format")
                        return [tuple(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
            except Exception as fallback_error:
                logger.error(f"Fallback also failed: {fallback_error}")
            
            return None

    def import_excel(
        self,
        file_content: bytes,
        filename: str,
        import_batch_id: Optional[str] = None,
        default_hourly_rate: Optional[Decimal] = None,
        default_markup: Optional[Decimal] = None,
        carry_forward_rates: bool = True,
    ) -> dict:
        """
        Import contract labor entries from an Excel file.
        
        Args:
            file_content: Excel file content as bytes
            filename: Original filename
            import_batch_id: Optional batch ID (generated if not provided)
            default_hourly_rate: Default rate to apply
            default_markup: Default markup to apply
            carry_forward_rates: Whether to use previous rates for vendors
            
        Returns:
            Dict with import results
        """
        if not import_batch_id:
            import_batch_id = str(uuid.uuid4())[:8]
        
        results = {
            "import_batch_id": import_batch_id,
            "total_rows": 0,
            "imported_count": 0,
            "skipped_count": 0,
            "error_count": 0,
            "errors": [],
            "skipped": [],  # Track why rows were skipped
            "unmatched_vendors": [],
            "unmatched_projects": [],
        }
        
        try:
            # Detect file type and load appropriately
            rows_iterator = self._load_excel_rows(file_content, filename)
            if rows_iterator is None:
                results["errors"].append({"row": 0, "error": "Could not load Excel file"})
                return results
            
            # Process rows (skip header row)
            row_num = 0
            for row in rows_iterator:
                row_num += 1
                results["total_rows"] += 1
                
                try:
                    # Skip empty rows
                    if not row or not any(row):
                        results["skipped_count"] += 1
                        results["skipped"].append({
                            "row": row_num + 1,
                            "reason": "Empty row"
                        })
                        continue
                    
                    # Parse row
                    parsed, skip_reason = self._parse_row(row, row_num)
                    if not parsed:
                        results["skipped_count"] += 1
                        results["skipped"].append({
                            "row": row_num + 1,
                            "reason": skip_reason or "Invalid data"
                        })
                        continue
                    
                    # Store raw names - vendor/project matching happens during review step
                    vendor_name = parsed.get("vendor_name", "")
                    job_name = parsed.get("job_name", "")
                    
                    # Calculate billing period from work date
                    work_date = parsed.get("work_date")
                    billing_period_start = ContractLabor.calculate_billing_period_start(work_date)
                    
                    # Create contract labor record with raw data
                    # Vendor, project, rates, and sub_cost_code assigned during review
                    contract_labor = self.repo.create(
                        vendor_id=None,  # Assigned during review
                        project_id=None,  # Assigned during review
                        employee_name=vendor_name,  # Raw name from Excel
                        work_date=work_date,
                        time_in=parsed.get("time_in"),
                        time_out=parsed.get("time_out"),
                        break_time=parsed.get("break_time"),
                        regular_hours=parsed.get("regular_hours"),
                        overtime_hours=parsed.get("overtime_hours"),
                        total_hours=parsed.get("total_hours"),
                        hourly_rate=None,  # Assigned during review
                        markup=None,  # Assigned during review
                        total_amount=None,  # Calculated during review
                        sub_cost_code_id=None,  # Assigned during review
                        description=parsed.get("notes"),
                        billing_period_start=billing_period_start,
                        status="pending_review",
                        import_batch_id=import_batch_id,
                        source_file=filename,
                        source_row=row_num + 1,  # Excel row number (1-based + header)
                        job_name=job_name,  # Raw job name from Excel
                    )
                    
                    results["imported_count"] += 1
                    
                except Exception as e:
                    logger.error(f"Error processing row {row_num + 1}: {e}")
                    results["errors"].append({
                        "row": row_num + 1,
                        "error": str(e)
                    })
                    results["error_count"] += 1
            
        except Exception as e:
            logger.exception(f"Error importing Excel file: {e}")
            results["errors"].append({"row": 0, "error": f"Failed to process Excel file: {str(e)}"})
            results["error_count"] += 1
        
        return results

    def _parse_row(self, row: tuple, row_num: int) -> Tuple[Optional[dict], Optional[str]]:
        """
        Parse a row from the Excel file.
        
        Returns:
            Tuple of (parsed_data, skip_reason)
            - parsed_data: Dict with parsed values or None if row should be skipped
            - skip_reason: Reason for skipping (None if not skipped)
        """
        # Get cell values safely
        def get_cell(index: int) -> Optional[str]:
            if index < len(row) and row[index] is not None:
                return str(row[index]).strip()
            return None
        
        # Parse date
        date_val = row[self.COL_DATE] if self.COL_DATE < len(row) else None
        work_date = self._parse_date(date_val)
        
        if not work_date:
            logger.debug(f"Row {row_num}: Skipping row with no valid date")
            return None, f"No valid date (value: '{date_val}')"
        
        # Get job and vendor name
        job_name = get_cell(self.COL_JOB)
        vendor_name = get_cell(self.COL_NAME)
        
        if not vendor_name:
            logger.debug(f"Row {row_num}: Skipping row with no vendor name")
            return None, "No vendor/worker name in column C"
        
        # Parse time values
        time_in = get_cell(self.COL_TIME_IN)
        time_out = get_cell(self.COL_TIME_OUT)
        break_time = get_cell(self.COL_BREAK_TIME)
        
        # Parse hours
        regular_hours = self._parse_time_to_hours(get_cell(self.COL_REGULAR_TIME))
        overtime_hours = self._parse_time_to_hours(get_cell(self.COL_OT))
        total_hours = self._parse_time_to_hours(get_cell(self.COL_TOTAL_TIME))
        
        # If total hours is not available, try to calculate from regular + OT
        if total_hours is None:
            if regular_hours is not None:
                total_hours = regular_hours
                if overtime_hours is not None:
                    total_hours = regular_hours + overtime_hours
        
        if total_hours is None or total_hours <= 0:
            logger.debug(f"Row {row_num}: Skipping row with no valid hours")
            return None, f"No valid hours (regular: {get_cell(self.COL_REGULAR_TIME)}, total: {get_cell(self.COL_TOTAL_TIME)})"
        
        # Get notes
        notes = get_cell(self.COL_NOTES)
        
        return {
            "work_date": work_date,
            "job_name": job_name,
            "vendor_name": vendor_name,
            "time_in": time_in,
            "time_out": time_out,
            "break_time": break_time,
            "regular_hours": regular_hours,
            "overtime_hours": overtime_hours,
            "total_hours": total_hours,
            "notes": notes,
        }, None  # No skip reason - row parsed successfully

    def _parse_date(self, date_val) -> Optional[str]:
        """
        Parse date value from Excel to YYYY-MM-DD format.
        
        Handles:
        - datetime objects
        - Strings like "Tuesday, January 20, 2026"
        - Strings like "2026-01-20"
        """
        if date_val is None:
            return None
        
        # Handle datetime objects
        if isinstance(date_val, datetime):
            return date_val.strftime("%Y-%m-%d")
        
        date_str = str(date_val).strip()
        if not date_str:
            return None
        
        # Try common formats
        formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%A, %B %d, %Y",  # "Tuesday, January 20, 2026"
            "%B %d, %Y",      # "January 20, 2026"
        ]
        
        for fmt in formats:
            try:
                parsed = datetime.strptime(date_str, fmt)
                return parsed.strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        # Try to extract date from string with regex
        # Pattern for "Month Day, Year"
        match = re.search(r"(\w+)\s+(\d{1,2}),?\s+(\d{4})", date_str)
        if match:
            month_name, day, year = match.groups()
            try:
                parsed = datetime.strptime(f"{month_name} {day}, {year}", "%B %d, %Y")
                return parsed.strftime("%Y-%m-%d")
            except ValueError:
                pass
        
        logger.warning(f"Could not parse date: {date_str}")
        return None

    def _parse_time_to_hours(self, time_str: Optional[str]) -> Optional[Decimal]:
        """
        Parse time string to decimal hours.
        
        Handles:
        - "08:30" -> 8.5
        - "02:45" -> 2.75
        - "8.5" -> 8.5 (already decimal)
        - "01:00" -> 1.0
        """
        if not time_str:
            return None
        
        time_str = time_str.strip()
        if not time_str:
            return None
        
        # Try HH:MM format
        if ":" in time_str:
            try:
                parts = time_str.split(":")
                hours = int(parts[0])
                minutes = int(parts[1]) if len(parts) > 1 else 0
                return Decimal(str(hours)) + Decimal(str(minutes)) / Decimal("60")
            except (ValueError, IndexError):
                pass
        
        # Try decimal format
        try:
            return Decimal(time_str)
        except InvalidOperation:
            pass
        
        logger.warning(f"Could not parse time: {time_str}")
        return None

    def _load_caches(self):
        """Load all vendors and projects into cache for faster lookups."""
        # Load vendors
        all_vendors = self.vendor_service.read_all()
        for vendor in all_vendors:
            if vendor.name:
                # Store by lowercase name for case-insensitive matching
                self._vendor_cache[vendor.name.lower()] = vendor
                # Also store by first word / abbreviation if available
                if vendor.abbreviation:
                    self._vendor_cache[vendor.abbreviation.lower()] = vendor
        
        # Load projects
        all_projects = self.project_service.read_all()
        for project in all_projects:
            if project.name:
                self._project_cache[project.name.lower()] = project
            if project.abbreviation:
                self._project_cache[project.abbreviation.lower()] = project

    def _match_vendor(self, vendor_name: str) -> Optional[Any]:
        """
        Match vendor by name.
        
        Tries:
        1. Exact match (case-insensitive)
        2. Partial match on first/last name
        """
        if not vendor_name:
            return None
        
        name_lower = vendor_name.lower().strip()
        
        # Exact match
        if name_lower in self._vendor_cache:
            return self._vendor_cache[name_lower]
        
        # Try matching by parts (first name, last name)
        for cached_name, vendor in self._vendor_cache.items():
            # Check if vendor name contains the search term
            if name_lower in cached_name or cached_name in name_lower:
                return vendor
        
        # Fall back to database lookup
        vendor = self.vendor_service.read_by_name(name=vendor_name)
        if vendor:
            self._vendor_cache[name_lower] = vendor
            return vendor
        
        return None

    def _match_project(self, job_name: str) -> Optional[Any]:
        """
        Match project by job name from Excel.
        
        Job format examples:
        - "HP - 6135 Hillsboro Pike"
        - "MR2 - 1577 Moran Rd"
        - "General"
        
        Matching strategy:
        1. Extract abbreviation if format is "ABBR - Name"
        2. Try exact match on name
        3. Try match on abbreviation
        4. Try partial match
        """
        if not job_name:
            return None
        
        job_lower = job_name.lower().strip()
        
        # Exact match on full job name
        if job_lower in self._project_cache:
            return self._project_cache[job_lower]
        
        # Try to extract abbreviation (format: "ABBR - Name" or "ABBR- Name")
        abbrev_match = re.match(r"^([A-Za-z0-9]+)\s*[-–]\s*(.+)$", job_name)
        if abbrev_match:
            abbreviation = abbrev_match.group(1).lower()
            # Try abbreviation match
            if abbreviation in self._project_cache:
                return self._project_cache[abbreviation]
        
        # Try partial matching
        for cached_key, project in self._project_cache.items():
            # Check if project name or abbreviation is in the job name
            if cached_key in job_lower or job_lower in cached_key:
                return project
        
        # Fall back to database lookup by name
        project = self.project_service.read_by_name(name=job_name)
        if project:
            self._project_cache[job_lower] = project
            return project
        
        return None

    def _get_rate_for_vendor(self, vendor_id: int) -> Tuple[Optional[Decimal], Optional[Decimal]]:
        """
        Get the carry-forward rate and markup for a vendor.
        First checks cache, then database.
        """
        if vendor_id in self._rate_cache:
            return self._rate_cache[vendor_id]
        
        hourly_rate, markup = self.repo.get_last_rate_for_vendor(vendor_id)
        self._rate_cache[vendor_id] = (hourly_rate, markup)
        return (hourly_rate, markup)

    def get_import_preview(
        self,
        file_content: bytes,
        max_rows: int = 10,
    ) -> dict:
        """
        Get a preview of what would be imported from an Excel file.
        Useful for showing the user what will be imported before committing.
        
        Args:
            file_content: Excel file content
            max_rows: Maximum number of rows to preview
            
        Returns:
            Dict with preview data
        """
        preview = {
            "total_rows": 0,
            "preview_rows": [],
            "vendors_found": [],
            "vendors_missing": [],
            "projects_found": [],
            "projects_missing": [],
        }
        
        try:
            workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
            sheet = workbook.active
            
            if not sheet:
                return preview
            
            self._load_caches()
            
            vendors_found = set()
            vendors_missing = set()
            projects_found = set()
            projects_missing = set()
            
            row_num = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_num += 1
                
                if not row or not any(row):
                    continue
                
                preview["total_rows"] += 1
                
                parsed = self._parse_row(row, row_num)
                if not parsed:
                    continue
                
                vendor_name = parsed.get("vendor_name", "")
                job_name = parsed.get("job_name", "")
                
                vendor = self._match_vendor(vendor_name)
                project = self._match_project(job_name)
                
                if vendor:
                    vendors_found.add(vendor_name)
                elif vendor_name:
                    vendors_missing.add(vendor_name)
                
                if project:
                    projects_found.add(job_name)
                elif job_name:
                    projects_missing.add(job_name)
                
                if len(preview["preview_rows"]) < max_rows:
                    preview["preview_rows"].append({
                        "row": row_num + 1,
                        "date": parsed.get("work_date"),
                        "job": job_name,
                        "vendor": vendor_name,
                        "hours": float(parsed.get("total_hours", 0)),
                        "notes": parsed.get("notes"),
                        "vendor_matched": vendor is not None,
                        "project_matched": project is not None,
                    })
            
            preview["vendors_found"] = list(vendors_found)
            preview["vendors_missing"] = list(vendors_missing)
            preview["projects_found"] = list(projects_found)
            preview["projects_missing"] = list(projects_missing)
            
            workbook.close()
            
        except Exception as e:
            logger.exception(f"Error generating preview: {e}")
        
        return preview
